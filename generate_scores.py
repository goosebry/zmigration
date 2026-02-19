"""
Generate Excel file with migration score predictions using 3 models
"""

import numpy as np
import csv

# Data from screenshot
data = [
    ("Dotty", 35, 149, 111, 50, 26),
    ("Korvath", 32, 83, 70, 28, 14),
    ("Officer Nolan", 30, 81, 65, 27, 14),
    ("Haram", 35, 144, 100, 41, 26),
    ("Jhayyy", 31, 93, 68, 26, 15),
    ("Lexsama", 32, 92, 80, 31, 18),
    ("DaddyC", 31, 82, 74, 32, 20),
    ("Bon/Axel", 30, 63, 65, 22, 13),
    ("Burn/Puppy Kisses", 30, 77, 68, 25, 13),
    ("Pala", 30, 77, 68, 21, 13),
    ("Bodach Glas", 30, 76, 54, 17, 12),
    ("Slamming Clams", 30, 50, 60, 18, 12),
    ("Roughneck", 30, 66, 53, 21, 9),
    ("Boozy", 30, 75, 58, 23, 12),
    ("Toilet Emoji", 30, 59, 60, 20, 9),
    ("Haagendaz", 31, 64, 69, 25, 11),
    ("Big Chungus", 30, 54, 51, 19, 9),
    ("M3RLin", 29, 60, 49, 18, 7.5),
    ("Supreme Leader", 31, 58, 58, 26, 7.7),
    ("Krys", 29, 59, 53, 15, 9.5),
    ("Lucy of Bane", 28, 52, 37, 14, 7),
    ("Qabhoes", 29, 46, 42, 15, 7),
    ("TeMu Tip", 29, 46, 44, 14, 7),
    ("Miss Behave", 29, 50, 50, 14, 7),
    ("Kain", 28, 54, 49, 20, 8),
    ("Yoona", 30, 48, 43, 15, 7),
    ("Azrathon", 27, 41, 47, 11, 8),
    ("Niooo", 31, 57, 48, 14, 8),
    ("Lai158", 30, 52, 50, 15, 10),
    ("Jenny TIC", 29, 44, 49, 15, 9),
    ("Notyy", 30, 41, 49, 11, 10),
    ("Dix F2P", 30, 44, 55, 18, 9),
    ("Boooger Bleed", 30, 39, 44, 14, 7),
    ("Macculate", 30, 46, 55, 16, 7),
    ("Max Lubricant", 30, 52, 42, 16, 7),
    ("Elio Six", 30, 48, 33, 13, 5.5),
    ("Joke Bear", 30, 54, 40, 14, 6),
    ("ez-CoffeeBean", 30, 52, 40, 13, 6),
    ("Sabina885", 30, 58, 39, 13, 6),
    ("Ezdie", 30, 55, 33, 14, 6),
    ("05Bastian07", 30, 47, 40, 16, 6.5),
    ("Azra-L", 30, 43, 45, 16, 9),
    ("Topgun 262", 31, 42, 35, 14, 6),
    ("YuNing", 30, 48, 44, 17, 7.5),
    ("V-Power", 31, 59, 55, 18, 10.5),
    ("Maxsama", 31, 50, 47, 18, 8),
    ("D D", 30, 39, 41, 14, 7),
    ("Armi", 30, 50, 45, 14, 7),
    ("Tomato Mama", 29, 45, 48, 16, 7),
    ("Ares98", 28, 43, 35, 12, 5),
    ("Cole88", 28, 28, 34, 13, 5),
    ("halowasmyidea", 29, 30, 34, 14, 6),
    ("Pitchie Kawaii", 29, 30, 34, 8, 6),
    ("Twinkle Twinkle", 28, 37, 36, 10, 5),
    ("Bird177", 27, 38, 28, 7, 5),
    ("Barbiegirl", 28, 37, 25, 5, 5),
    ("QQ123456", 27, 31, 26, 9, 4),
    ("Asian Letters Ohno", 28, 38, 29, 11, 5),
    ("Ez Moon", 29, 36, 30, 10, 4.5),
    ("ZaraA", 28, 40, 30, 11, 5),
    ("Son With Us", 28, 32, 32, 8, 6),
]

def model5_power_law(total):
    """Score = 0.0021 × Total^2.0141"""
    return 0.0021 * (total ** 2.0141)

def model6_weighted(S, H, T, V):
    """Base = 0.20×S + 0.70×H + 0.10×T + 0.30×V; Score = 0.0092 × Base^2.0886"""
    base = 0.200 * S + 0.700 * H + 0.100 * T + 0.300 * V
    return 0.0092 * (base ** 2.0886)

def model3_scaling(total, est_ratio=0.26):
    """Est × (1.2734 × ln(Total) - 5.0378)"""
    est = total * est_ratio
    scaling = 1.2734 * np.log(total) - 5.0378
    return est * scaling

# Generate predictions
results = []
for row in data:
    name, hq, structure, hero, tech, truck = row
    total = structure + hero + tech + truck
    
    pred_m5 = model5_power_law(total)
    pred_m6 = model6_weighted(structure, hero, tech, truck)
    pred_m3 = model3_scaling(total)
    avg = (pred_m5 + pred_m6 + pred_m3) / 3
    
    results.append({
        'Name': name,
        'HQ': hq,
        'Structure_M': structure,
        'Hero_M': hero,
        'Tech_M': tech,
        'Truck_M': truck,
        'Total_M': total,
        'Model5_PowerLaw': round(pred_m5, 1),
        'Model6_Weighted': round(pred_m6, 1),
        'Model3_Scaling': round(pred_m3, 1),
        'Average': round(avg, 1),
    })

# Write to CSV (can be opened in Excel)
csv_path = '/Users/garethberry/zmigration/migration_scores.csv'
with open(csv_path, 'w', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=results[0].keys())
    writer.writeheader()
    writer.writerows(results)

print(f"CSV file written to: {csv_path}")
print(f"\nTotal players processed: {len(results)}")

# Also try to write Excel format if openpyxl is available
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Migration Scores"
    
    # Headers
    headers = list(results[0].keys())
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, row_data in enumerate(results, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data[key])
    
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Wider name column
    ws.column_dimensions['A'].width = 20
    
    xlsx_path = '/Users/garethberry/zmigration/migration_scores.xlsx'
    wb.save(xlsx_path)
    print(f"Excel file written to: {xlsx_path}")
    
except ImportError:
    print("\nNote: openpyxl not installed. CSV file created instead.")
    print("To install: pip3 install openpyxl")
    print("The CSV file can be opened directly in Excel.")

# Print summary table
print("\n" + "=" * 100)
print("MIGRATION SCORE PREDICTIONS")
print("=" * 100)
print(f"{'Name':<20} {'HQ':>3} {'Total':>7} {'Model5':>8} {'Model6':>8} {'Model3':>8} {'Average':>8}")
print("-" * 100)
for r in results:
    print(f"{r['Name']:<20} {r['HQ']:>3} {r['Total_M']:>7.1f} {r['Model5_PowerLaw']:>8.1f} {r['Model6_Weighted']:>8.1f} {r['Model3_Scaling']:>8.1f} {r['Average']:>8.1f}")
